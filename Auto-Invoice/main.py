import tkinter as tk
import openpyxl
import inflect
import shutil

root = tk.Tk()
root.title("Auto Invoice Generator - 1.0.4")
root.geometry("640x640")

start_point = tk.IntVar()
stop_point = tk.IntVar()

usd_rate = tk.IntVar()
au_rate = tk.IntVar()
euro_rate = tk.IntVar()
gbp_rate = tk.IntVar()

output_file_path = tk.StringVar()


my_file = open('data.txt', 'r')
last_invoice = my_file.read()
my_file.close()

new_i_data = "0"



def run_code():
    new_start = start_point.get()
    new_stop = stop_point.get()
    new_output_file_path = output_file_path.get()
    new_usd = usd_rate.get()
    new_au = au_rate.get()
    new_euro = euro_rate.get()
    new_stops = new_stop + 2
    new_starts = new_start + 1
    new_gbp = gbp_rate.get()
    global new_i_data


    def num2word(num):
        p = inflect.engine()

        words = p.number_to_words(num).capitalize()
        return words

    def remove_char(str, chars):
        return "".join(c for c in str if c not in chars)

    file = openpyxl.load_workbook("rawdata.xlsx")

    sheet = file['Sheet1']

    out_workbook = openpyxl.load_workbook("Invoice_or.xlsx")
    out_sheet = out_workbook["Sheet1"]

    for val in range(new_starts, new_stops):

        # fetching the data.
        cell_value_for_id = sheet.cell(val, 1).value
        cell_value_for_name = sheet.cell(val, 2).value
        cell_value_for_add1 = sheet.cell(val, 3).value
        cell_value_for_add2 = sheet.cell(val, 4).value
        cell_value_for_city = sheet.cell(val, 5).value
        cell_value_for_state = sheet.cell(val, 6).value
        cell_value_for_zip = sheet.cell(val, 7).value
        cell_value_for_country = sheet.cell(val, 8).value
        cell_value_for_mail = sheet.cell(val, 9).value
        cell_value_for_custom_label = sheet.cell(val, 10).value
        cell_value_for_currency = sheet.cell(val, 13).value
        get_curr = str(sheet.cell(val, 13).value)

        cell_value_for_total = sheet.cell(val, 13).value
        cell_value_for_paypal = sheet.cell(val, 14).value
        # cell_value_for_title = sheet.cell(val, )
        cell_value_for_currency_str = str(cell_value_for_currency)
        cell_value_total_for_words = cell_value_for_total

        if cell_value_for_add2 is None:
            cell_value_for_add2 = "--"

        if cell_value_for_zip is None:
            cell_value_for_zip = "--"

        if str(cell_value_for_custom_label) == "SESA100":

            new_title = "Sesa Hair Oil 100ml"

            new_quantity = 1

        elif str(cell_value_for_custom_label) == "HARITAKI1KG":

            new_title = "MB Herbals Haritaki Powder 100g"
            new_quantity = 10

        elif str(cell_value_for_custom_label) == "ONIONOIL200":
            new_title = "MB Herbals Onion Hair Oil 200ml"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "ALUM10G":

            new_title = "MB Herbals Alum Powder 10g "
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "HARITAKI300G":

            new_title = "MB Herbals Haritaki Powder 100g"
            new_quantity = 3

        elif str(cell_value_for_custom_label) == "BRAHMI100":

            new_title = "MB Herbals Brahmi Powder 100g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "ALUM227G":

            new_title = "MB Herbals Alum Powder  227g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "INDIGO100G":

            new_title = "MB Herbals Indigo Powder 100g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "INDIGO227G":

            new_title = "MB Herbals Indigo Powder 227g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "SLANCETS200":

            new_title = "Accu Chek SoftClix Lancets 200"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "MAHABH100":
            new_title = "MahaBhringaraj Oil 100ml"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "SHK227G":

            new_title = "MB Herbals Shikakai Powder 227g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "MAHABH200":

            new_title = "MahaBhringaraj Oil 200ml"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "BORIC500":

            new_title = "Boric Acid Powder 100g"
            new_quantity = 5

        elif str(cell_value_for_custom_label) == "R83":

            new_title = "Dr. Reckeweg Homoeppathy Drops R83"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "R87":

            new_title = "Dr. Reckeweg Homoeppathy Drops R87"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "HIBISCUS10G":

            new_title = "MB Herbals Hibiscus Powder 10g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "FAGONIA500":

            new_title = "MB Herbals Fagonia cretica Powder 500g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "SOFTCLIXDEV":

            new_title = "Accu Chek SoftClix Lancing Device"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "BRAHMI227":

            new_quantity = 1
            new_title = "MB Herbals Brahmi Powder 227g"

        elif str(cell_value_for_custom_label) == "HARITAKI100G":

            new_title = "MB Herbals Haritaki Powder 100g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "VT15G":

            new_title = "Vicco Turmeric Cream 15g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "MAHABH300":

            new_title = "MahaBhringaraj Oil 300ml"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "MAHABH100":

            new_title = "MahaBhringaraj Oil 100ml"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "FULLERS100":

            new_title = "MB Herbals Fullers Earth 100g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "ActiveKit":

            new_title = "Accu Chek Active Glucometer Kit"
            new_quantity = 1



        elif str(cell_value_for_custom_label) == "HARITAKI50G":

            new_title = "MB Herbals Haritaki Powder 50g"
            new_quantity = 1

        elif str(cell_value_for_custom_label) == "AKARKARA100":

            new_title = "MB Herbals Akarkara Powder 100g"
            new_quantity = 1

        # elif str(cell_value_for_custom_label) == ""

        else:
            new_title = "Sku data not found."

            new_quantity = "null"

        # write data
        write_buyer_id = out_sheet.cell(6, 7)
        write_buyer_id.value = cell_value_for_id

        write_buyer_name = out_sheet.cell(13, 2)
        write_buyer_name.value = cell_value_for_name

        write_buyer_add1 = out_sheet.cell(14, 2)
        write_buyer_add1.value = cell_value_for_add1

        write_buyer_add2 = out_sheet.cell(15, 2)
        write_buyer_add2.value = cell_value_for_add2

        write_buyer_city = out_sheet.cell(16, 2)
        write_buyer_city.value = cell_value_for_city

        write_buyer_state = out_sheet.cell(17, 2)
        write_buyer_state.value = cell_value_for_state

        write_buyer_zip = out_sheet.cell(18, 2)
        write_buyer_zip.value = cell_value_for_zip

        write_buyer_country = out_sheet.cell(19, 2)
        write_buyer_country.value = cell_value_for_country

        write_buyer_mail = out_sheet.cell(20, 2)
        write_buyer_mail.value = cell_value_for_mail

        write_buyer_title = out_sheet.cell(24, 2)
        write_buyer_title.value = new_title

        write_buyer_quantity = out_sheet.cell(24, 4)
        write_buyer_quantity.value = new_quantity

        cell_value_for_total_new = str(cell_value_for_total)
        cell_value_for_total_new_2 = remove_char(cell_value_for_total_new, "$AUGBP")

        filter_total = float(cell_value_for_total_new_2)

        # print(filter_total)
        is_au = "AU" in get_curr

        is_euro = "EU" in get_curr

        is_gbp = "GBP" in get_curr

        if is_au is True:

            new_amount = filter_total * new_au


        elif is_euro is True:

            new_amount = filter_total * new_euro

        elif is_gbp is True:

            new_amount = filter_total * new_gbp


        else:
            new_amount = filter_total * new_usd

        # cell_value_for_currency_str_new = remove_char(cell_value_for_currency_str, ".12345$6789")

        write_buyer_curr = out_sheet.cell(24, 5)
        write_buyer_curr.value = "INR"

        new_amount_round = round(new_amount)

        write_buyer_total = out_sheet.cell(42, 9)
        write_buyer_total.value = new_amount_round

        write_buyer_id = out_sheet.cell(9, 7)
        write_buyer_id.value = cell_value_for_paypal
        only = " only."
        word_string = num2word(new_amount_round)
        word_string_2 = f"{word_string} {only}"
        write_buyer_total_in_words = out_sheet.cell(39, 2)
        write_buyer_total_in_words.value = word_string_2

        out_workbook.save("Invoice_or.xlsx")
        # create new output file.

        invoice_no = str(sheet.cell(val, 1).value)

        wb = openpyxl.Workbook()

        # Get workbook active sheet
        # from the active attribute
        sheet1 = wb.active

        # One can change the name of the title
        sheet1.title = "Invoice"
        if new_title == "Sku data not found.":

            wb.save(f"{new_output_file_path}Invoice_{invoice_no} NO SKU.xlsx")
            shutil.copyfile("Invoice_or.xlsx",
                            f"{new_output_file_path}Invoice_{invoice_no} NO SKU.xlsx")
        else:
            wb.save(f"{new_output_file_path}Invoice_{invoice_no}.xlsx")
            shutil.copyfile("Invoice_or.xlsx",
                            f"{new_output_file_path}Invoice_{invoice_no}.xlsx")


        new_i_data = invoice_no


        if val == new_stop + 1:


            new_file_i = open("data.txt", "w")
            new_file_i.write(new_i_data)
            new_file_i.close()

        job_done = tk.Label(root, text="JOB DONE!!", font=("arial", 20, "bold"), fg="darkgreen").place(x=180, y=515)

        out_workbook.save("Invoice_or.xlsx")
        out_workbook.close()


heading = tk.Label(root, text="M.B Herbals", font=("arial", 50, "bold"), fg="brown").pack()

instruct = tk.Label(root,
                    text=f"Please paste 'rawdata.xlsx' before running program.\nIn case of error rerun the software.\nLast invoice generated was {last_invoice}",
                    font=("arial", 15),
                    fg="black").place(x=10, y=90)
label_for_start = tk.Label(root, text="Generate invoice from: ", font=("arial", 15, "bold"), fg="black").place(x=10,
                                                                                                               y=215)
entry_box_start_number = tk.Entry(root, textvariable=start_point, width=25, bg="lightgreen").place(x=280, y=215)

label_for_stop = tk.Label(root, text="Generate invoice to: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=250)

label_for_usd_rate = tk.Label(root, text="USD rate: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=290)
entry_box_usd_rate = tk.Entry(root, textvariable=usd_rate, width=25, bg="lightgreen").place(x=278, y=290)

label_for_au_rate = tk.Label(root, text="Au rate: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=325)
entry_box_au_rate = tk.Entry(root, textvariable=au_rate, width=25, bg="lightgreen").place(x=278, y=325)

label_for_eur_rate = tk.Label(root, text="Euro rate: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=360)
entry_box_eur_rate = tk.Entry(root, textvariable=euro_rate, width=25, bg="lightgreen").place(x=278, y=360)

label_for_gbp_rate = tk.Label(root, text="GBP rate: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=395)
entry_box_gbp_rate = tk.Entry(root, textvariable=gbp_rate, width=25, bg="lightgreen").place(x=278, y=395)
label_for_path = tk.Label(root, text="Output file path: ", font=("arial", 15, "bold"), fg="black").place(x=10, y=425)
entry_box_path = tk.Entry(root, textvariable=output_file_path, width=25, bg="lightgreen").place(x=278, y=430)

entry_box_stop_number = tk.Entry(root, textvariable=stop_point, width=25, bg="lightgreen").place(x=278, y=255)
save_stop_num = tk.Button(root, text="Generate Invoices", width=25, bg="lightblue", command=run_code).place(x=275,
                                                                                                            y=480)

made_by = tk.Label(root, text="Author: Azim Mustufa Baldiwala.", font=("arial", "15", "bold"), fg="red").place(x=300,
                                                                                                               y=550)

root.mainloop()
